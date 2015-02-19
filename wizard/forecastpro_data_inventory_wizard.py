from openerp.osv import fields, osv, orm
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
import base64
from openerp.tools.translate import _

class forecastpro_data_inventory_wizard(osv.osv_memory):
    _name = "forecastpro.data.inventory.wizard"
    _description = "Load inventory data spreadsheet file to Forecast Pro"
    _columns = {
        'year': fields.integer('Year'),
        'file': fields.binary('Select file')
    }

    def load_file(self, cr, uid, ids, context=None):
        if context is None:
            context = {}
        form_data = self.read(cr, uid, ids, ['file', 'year'])[0]
        branch_pool = self.pool.get('res.partner')
        product_pool = self.pool.get('product.product')
        data_pool = self.pool.get('forecastpro.data.inventory')
        year = form_data['year']
        xlsfile = base64.decodestring(form_data['file'])
        fileobj = NamedTemporaryFile('w+b', delete=False)
        fileobj.write(xlsfile)
        fileobj.close()
        lineasCorrectas = []
        lineasError = []
        numeroLinea = 0
        wb = load_workbook(fileobj.name)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        for row in ws.iter_rows():
            numeroLinea += 1
            if not (row[0].data_type == 's' and row[0].value.lower() == 'branch'):
                mensajeError = ''
                branch_id = 0
                branch_ids = branch_pool.search(cr, uid, [('active', '=', True),('customer', '=', True),('ref', '=', row[0].value)])
                if branch_ids:
                    for branch in branch_pool.browse(cr, uid, branch_ids, context=context):
                        branch_id = branch.id
                else:
                    if mensajeError != '':
                        mensajeError += ', '
                    mensajeError += 'Invalid branch code (' + str(row[0].value) + ')'
                product_id = 0
                product_ids = product_pool.search(cr, uid, [('active', '=', True),('ean13', '=', row[1].value)])
                if product_ids:
                    for product in product_pool.browse(cr, uid, product_ids, context=context):
                        product_id = product.id
                else:
                    if mensajeError != '':
                        mensajeError += ', '
                    mensajeError += 'Invalid isbn (' + str(row[1].value) + ')'
                if not row[2].value.lower() in ['inventory on hand', 'inventory on order']:
                    if mensajeError != '':
                        mensajeError += ', '
                    mensajeError += 'Invalid inventory type (' + str(row[2].value) + ')'
                if mensajeError == '':
                    lineasCorrectas.append(
                        {
                            'row': row,
                            'branch_id': branch_id,
                            'product_id': product_id,
                        }
                    )
                else:
                    lineasError.append(str(numeroLinea) + ': ' + mensajeError)
        if len(lineasError) > 0:
            mensajeError = 'Errors found in spreadsheet row'
            if len(lineasError) > 1:
                mensajeError += 's'
            mensajeError += ':\n'
            for linea in lineasError:
                mensajeError += linea + "\n"
            raise osv.except_osv(_('Error!'),_(mensajeError))
        else:
            for lineaInsertar in lineasCorrectas:
                row = lineaInsertar['row']
                rowLength = len(row) - 1
                inventoryType = 'onhand'
                if str(row[2].value).lower() == 'inventory on order':
                    inventoryType = 'onorder'
                data = {
                    'branch': lineaInsertar['branch_id'],
                    'product_id': lineaInsertar['product_id'],
                    'inventory_type': inventoryType
                }
                month = 1
                for count in range(3, 14):
                    cr.execute("""
                        Delete From forecastpro_data_inventory
                        Where branch = %s
                        And product_id = %s
                        And inventory_type = %s
                        And year = %s
                        And month = %s
                    """, (data['branch'], data['product_id'], inventoryType, year, month))
                    if row[count].data_type == 'n':
                        if row[count].value != 0:
                            data['year'] = year
                            data['month'] = month
                            data['units'] = row[count].value
                            data_pool.create(cr, uid, data, context=context)
                    month = month + 1
                    if month > 12:
                        month = 1
                        year = year + 1
        return {
            'name': _('Forecast inventory data'),
            'view_type': 'form',
            'view_mode': 'graph',
            'res_model': 'forecastpro.data.inventory',
            'type': 'ir.actions.act_window',
            'context': context,
        }