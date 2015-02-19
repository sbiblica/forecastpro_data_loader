from openerp.osv import fields, osv, orm
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
import base64
from openerp.tools.translate import _

class forecastpro_generate_data_file_wizard(osv.osv_memory):
    _name = "forecastpro.generate.data.file.wizard"
    _description = "Generate spreadsheet file"
    _columns = {
        'branch': fields.many2one('res.partner', 'Branch', required=True, select=True),
        'year': fields.integer('Year'),
        'name': fields.char('File Name', readonly=True),
        'data': fields.binary('File', readonly=True),
        'state': fields.selection([('choose', 'choose'),   # choose year
                                   ('get', 'get')])        # get the file
    }

    _defaults = {
        'state': 'choose',
        'year': 2014,
    }

    def generate_file(self, cr, uid, ids, context=None):
        if context is None:
            context = {}
        this = self.browse(cr, uid, ids)[0]
        form_data = self.read(cr, uid, ids, ['year', 'branch'])[0]
        branch = form_data['branch']
        year = form_data['year']

        fileobj = NamedTemporaryFile('w+b')
        xlsfile = fileobj.name
        fileobj.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "UBS FWA External Modifiers"
        ws['A1'].value = "Branch"
        ws['B1'].value = "ISBN"
        ws['C1'].value = "Modifier"

        cr.execute("""
            Select C.ref, P.ean13
            From forecastpro_data_sales F
            Join product_product P
            On P.id = F.product_id
            And P.ean13 Is Not Null
            Join res_partner C
            On C.id = F.branch
            And C.customer
            Where F.branch = %s
            And F.year = %s
            and F.Month = 1
            Order By P.ean13
        """, (branch[0],year,))
        row = 2
        for forecast_line in cr.dictfetchall():
            ws.cell(row=row, column=1).value = forecast_line['ref']
            ws.cell(row=row, column=2).value = forecast_line['ean13']
            ws.cell(row=row, column=3).value = '\UPPER=0 \LEADTIME=0.0'
            row += 1

        wb.save(filename=xlsfile)

        spreadsheet_file = open(xlsfile, "rb")
        binary_data = spreadsheet_file.read()
        spreadsheet_file.close()
        out = base64.b64encode(binary_data)
        self.write(cr, uid, ids, {
            'state': 'get',
            'name': "UBS_FWA_External_Modifiers.xlsx",
            'data': out
        }, context=context)
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'forecastpro.generate.data.file.wizard',
            'view_mode': 'form',
            'view_type': 'form',
            'res_id': this.id,
            'views': [(False, 'form')],
            'target': 'new',
        }
